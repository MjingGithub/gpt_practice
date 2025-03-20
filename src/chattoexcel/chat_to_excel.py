import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl import load_workbook
import panel as pn  # GUI
pn.extension()

panels = [] # collect display 
# 初始化上下文对象
pn.state.wb = None  # 
pn.state.system_prompt = None

def create_system_prompt(fileName):
    wb = load_workbook(filename = fileName)
    system_prompt="You are working with a Workbook in Python.The name of the workbook is `wb`."
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        header = [cell.value for cell in sheet[1]]
        max_rows = sheet.max_row
        system_prompt+=f"\nwb['{sheet_name}'] has {max_rows} rows with excel index and head(10) rows as below:\n"
        column_index=[]
        for i in range(len(header)):
            column_index.append(get_column_letter(i+1))
        data=[]
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        df = pd.DataFrame(data)
        df.columns = column_index
        system_prompt+=df.head(10).to_markdown(index=False)+"\n"

    print(system_prompt)
    return system_prompt


import sys
from pathlib import Path # if you haven't already done so
file = Path(__file__).resolve()
parent, root = file.parent, file.parents[1]
sys.path.append(str(root))

# Additionally remove the current file's directory from sys.path
try:
    sys.path.remove(str(parent))
except ValueError: # Already removed
    pass


from langchain_experimental.agents import create_pandas_dataframe_agent
# from langchain_experimental.agents.agent_toolkits.csv.base import create_csv_agent
from langchain_community.chat_models import ChatOpenAI
from langchain.agents.agent_types import AgentType

'''
    利用create_pandas_dataframe_agent根据指令生成openai函数执行
'''
# df = pd.read_excel('菜鸟demp.xlsx')
# agent = create_pandas_dataframe_agent(
#     ChatOpenAI(temperature=0, model="gpt-4o-mini"),
#     df,
#     verbose=True,
#     agent_type=AgentType.OPENAI_FUNCTIONS,
#     allow_dangerous_code=True
# )


# agent.invoke("求所有费用项的总金额")

# agent.invoke("修改费用项'hhh费'为'仓租费'，并保存到文件'菜鸟demp1.xlsx'")

# agent.invoke("读取上面对话输出的最新的文件，按费用项和计费币种汇总金额，原文件sheet不覆盖，新内容保存到一个新的sheet叫汇总，最终一起保存到文件'菜鸟demp1.xlsx'")


'''
    利用openpyxl根据执行生成对应函数执行
'''
from langchain_core.prompts import ChatPromptTemplate
from commonlib.lang_chain_util import invokeAIResponseFrom_langChain
# 提取代码部分
import re
from io import BytesIO
def extract_code(response_func_msg):
    code_snippet = re.search(r'```python\n(.*?)\n```', response_func_msg, re.DOTALL)
     # 定义一个字典用于保存局部变量
    local_vars = {}
    if code_snippet:
        extracted_code = code_snippet.group(1)
        print("提取出的 Python 代码如下：")
        print(extracted_code)
        final_code = extracted_code.replace("wb","pn.state.wb")
        print(f"提取出的 Python 代码如下final_code：{final_code}")
        exec(final_code,{"pn":pn},local_vars)
    else:
        print("未找到 Python 代码片段。")

    return local_vars

# 代码生成agent
def code_generate_agent(user_prompt):
      context= [{'role':'system', 'content':f"{pn.state.system_prompt}，根据openpyxl库生成以下用户提问的代码，剔除加载工作簿部分，直接使用我的上下文中输入的wb工作簿,如果有返回结果，结果赋值给我的全局变量_result"}]
      context.append({'role':'user', 'content':f"{user_prompt}"})
      response = invokeAIResponseFrom_langChain(context) 
      return response

# 代码纠正agent
def code_revise_agent(exec_code,error_msg):
      context= [{'role':'system', 'content':f"{pn.state.system_prompt}，你负责纠正基于以上wb表格执行出现的代码错误，请注意不要改变变量名称和代码执行意图"}]
      context.append({'role':'user', 'content':f"我执行的代码是：{exec_code}，错误信息是：{error_msg}，请帮我修正生成正确的代码"})
      response = invokeAIResponseFrom_langChain(context) 
      return response


# 处理提示词和响应
def collect_messages(_):
    if file_input.value is not None:
        prompt = inp.value_input
        inp.value = ''
        if prompt == '' or prompt is None:
            return
        # File
        response = code_generate_agent(prompt)
        try:
            execution_results = extract_code(response_func_msg=response)
        except Exception as e:
            print(f"execute python code error:{str(e)}")
            # 执行自动纠正
            revise_code_msg = code_revise_agent(response,str(e))
            print(f"修正后的代码是：{revise_code_msg}")
            execution_results = extract_code(response_func_msg=revise_code_msg)
            response = revise_code_msg

        print(f"execution_results:{execution_results}")
        if execution_results['_result'] is not None:
            final_response = f"执行代码：\n {response}\n 最终结果为：{execution_results['_result']}" 
        else:
            final_response = f"执行代码：\n {response} "
        print(f"final_response:{final_response}")
        context.append({'role':'assistant', 'content':f"{response}"})
        panels.append(
            pn.Row('User:', pn.pane.Markdown(prompt, width=600)))
        panels.append(
            pn.Row('Assistant:', pn.pane.Markdown(final_response, width=600, styles={'background-color': '#F6F6F6'})))
    
        return pn.Column(*panels)

#https://panel.holoviz.org/reference/widgets/FileInput.html
file_input = pn.widgets.FileInput(accept='.xlsx,.xls')
context=[]

# 创建显示区域
preview_pane = pn.pane.Markdown(None)

#用户上传文件回调
def process_file(event):
    if file_input.value is not None:
         # 读取Excel文件并保存到上下文对象 wb
        pn.state.wb = load_workbook(filename = file_input.filename)
        print(f"fileName:{file_input.filename}")
        file_input.save(file_input.filename)
        pn.state.system_prompt = create_system_prompt(file_input.filename)
        preview_pane.object =  pn.state.system_prompt
     
# 绑定事件
file_input.param.watch(process_file, 'value')

#https://panel.holoviz.org/reference/widgets/TextInput.html#display
inp = pn.widgets.TextInput(placeholder='请输入您的问题...')
button_conversation = pn.widgets.Button(name="Click To Chat!",button_type='primary')

# 创建下载组件
def download_workbook():
    output = BytesIO()
    pn.state.wb.save(output)
    output.seek(0)
    return output

file_download = pn.widgets.FileDownload(
    callback=download_workbook,
    file='final_output.xlsx',
    button_type='primary',
    label='下载最终文件'
)


interactive_conversation = pn.bind(collect_messages, button_conversation)
dashboard = pn.Column(
    "Chat to Excel",
    pn.Row(file_input,file_download),
    preview_pane,
    pn.Row(inp,button_conversation),
    pn.panel(interactive_conversation, loading_indicator=True),
)

dashboard.servable()

    








